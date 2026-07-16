"""
SCED - Concise glossary of the produced statistics
==================================================

``explain(*keys)`` returns, as briefly as possible, *what each statistic
measures*, *how to read it* and *the pitfall* to be aware of. ``explain()``
without argument lists everything; ``explain_report(model_info)`` explains the
statistics present in a pipeline result dict.
"""

GLOSSARY = {
    # - Primary inference -
    "randomization test": "SCED primary inference: the condition assignment is re-drawn under the scheme that was ACTUALLY drawable and the observed statistic is compared to that distribution. Exact because randomization took place (no distributional assumption).",
    "p-value": "Proportion of permuted assignments as/more extreme than the observed one, p=(1+#)/(1+B). <0.05 = departure beyond assignment chance.",
    "partial F": "Partial F: share of variance explained by the CONDITION beyond the nuisance (time, unit). Larger = clearer effect. Omnibus (holds for 2 or k conditions).",
    "difference of means": "Two-condition statistic: mean(target) - mean(reference), in raw units. Two-sided test.",
    "omnibus variance": "Statistic for >=3 conditions: variance of the per-condition means. Large = more differentiated conditions. One-sided. IGNORES condition order.",
    "ordered trend": "Statistic for ORDERED conditions (dose-response, e.g. MI 4/5/6 s), IN ADDITION to the omnibus. OLS slope on dose (linear contrast) - change per step - or Jonckheere-Terpstra (ranks). The RT framework admits any statistic (Edgington & Onghena 2007); not validated for alternating SCED (Michiels 2017: only MD/NAP) -> pre-specify (Manolov 2021).",
    "jonckheere": "Jonckheere-Terpstra: non-parametric test of an ORDERED alternative (ranks; Terpstra 1952, Jonckheere 1954). Detects a monotone trend across ordered conditions without a linearity assumption.",
    "stratified test": "Group version: unit = BLOCK, WITHIN-unit permutation, COMMON effect tested by partial F. Between-patient differences are neutralized.",

    # - Permutation schemes -
    "draper-stoneman": "Scheme: the condition LABELS are permuted. Exact for a randomized design with condition independent of time; least robust under condition/time collinearity.",
    "freedman-lane": "Scheme: the RESIDUALS of the reduced model (Y~time) are permuted, not the labels. Robust under condition/time confounding; recommended default (Winkler 2014).",
    "condition-time R2": "Condition/time collinearity (R2 of time~condition). Low -> Draper-Stoneman; high (temporal imbalance) -> Freedman-Lane.",

    # - Detrend / trend diagnostic -
    "detrend": "time_covariate (formerly 'detrend'): puts the time slope AS A COVARIATE of the model (none / linear / log; auto = diagnostic suggestion) - not a separate pre-detrend. Prevents time from being confounded with condition; with FL, the slope is modeled within the test itself.",
    "mann-kendall": "Non-parametric test of a (monotone) temporal trend. p<0.05 = learning/drift present.",
    "hamed-rao": "Mann-Kendall correction for autocorrelation (otherwise too many false trend positives).",
    "sen's slope": "Robust slope (median of pairwise slopes): magnitude of the trend, per session.",
    "autocorrelation": "Lag-1 correlation / Durbin-Watson: dependence of residuals over time. Does NOT invalidate the randomization test, but informs detrending.",

    # - Non-overlap effect sizes -
    "tau-u": "Rank concordance treatment vs baseline, BOUNDED -1..1 (0=none). Baseline trend correction: Tarlow 2016 (Theil-Sen, 'adj', n_A>=7) or bounded denominator Brossard 2018 ('trend_a'). DO NOT use the n_a.n_b denominator of Parker 2011a (unbounded, values >1). Variant selected via the Fingerhut 2021 flowchart. Interpret via Critical Tau-U (Fingerhut) rather than the Vannest & Ninci 2015 thresholds.",
    "nap": "Non-overlap of All Pairs: % of (baseline x treatment) pairs where treatment wins. 0.5=chance; ~0.66 weak, ~0.92 medium, >0.92 strong.",
    "pnd": "% of treatment points beyond the BEST baseline point. Simple but biased (1 extreme point dominates) - prefer NAP/Tau-U.",
    "pem": "% of treatment points beyond the baseline MEDIAN. Less sensitive to extremes than PND.",

    # - Contrasts / standardized effect -
    "hedges g": "STANDARDIZED mean difference (small-sample corrected). 0.2/0.5/0.8 = small/medium/large. Read with its CI (excludes 0 = significant).",
    "pairwise contrasts": "Post-hoc: WHICH conditions differ pairwise. g + CI + permutation p, corrected Holm (FWER) and FDR (Benjamini-Hochberg).",
    "holm": "Multiplicity correction controlling the risk of >=1 false positive (FWER). Conservative, adjusted p >= raw p.",
    "fdr": "Benjamini-Hochberg correction: controls the expected proportion of false positives among the significant ones. Less conservative than Holm.",

    # - Heterogeneity (group) -
    "heterogeneity": "Does the effect VARY across patients? (unit x condition interaction). Complements the common effect, which can mask a non-responder or cancel out.",
    "interaction": "Partial F of the unit x condition interaction (Freedman-Lane, APPROXIMATE: the 'equal effects' null is not the strong null).",
    "cochran q": "Meta-analytic test of heterogeneity of the per-unit effects. Q p<0.05 = heterogeneous effects.",
    "i2": "% of the variance of the effects due to true heterogeneity (not chance). ~25/50/75 % = weak/medium/strong.",

    # - Power -
    "power": "Probability of detecting the effect (p<alpha), estimated by SIMULATION with the real test. Depends on effect/sd, n sessions, n units.",
    "mdes": "Smallest effect (Cohen's d) detectable at 80 % given the design. The correct retrospective read ('what could I have caught?').",
    "retrospective power": "Power of the realized design for a SPECIFIED effect (not the observed effect - the observed-power fallacy, Hoenig & Heisey 2001).",
}

# alias -> canonical key (entry tolerance)
_ALIASES = {
    "tau_u": "tau-u", "tauu": "tau-u", "tau": "tau-u",
    "freedman": "freedman-lane", "fl": "freedman-lane",
    "draper": "draper-stoneman", "ds": "draper-stoneman",
    "f": "partial F", "partial_f": "partial F", "randomization": "randomization test",
    "p": "p-value", "pvalue": "p-value", "sens slope": "sen's slope",
    "i2": "i2", "isquared": "i2", "q": "cochran q", "g": "hedges g",
    "minimum detectable effect": "mdes", "autocorr": "autocorrelation",
    "time_covariate": "detrend", "time covariate": "detrend",
    "i²": "i2",
}


def _normalise(key):
    """Lowercase/strip a key and resolve it through the alias table to its canonical form."""
    k = str(key).strip().lower()
    return _ALIASES.get(k, k)


def explain(*keys, as_dict=False):
    """Concise explanation of one or more statistics. Without argument, returns the
    whole glossary. ``as_dict=True`` returns a dict {key: text} instead of a
    print-ready text."""
    items = (list(GLOSSARY) if not keys
             else [_normalise(k) for k in keys])
    out = {}
    for k in items:
        out[k] = GLOSSARY.get(k, "(unknown - see explain() for the list)")
    if as_dict:
        return out
    return "\n".join(f"- {k} - {v}" for k, v in out.items())


def explain_report(model_info):
    """Explain the statistics detected in a pipeline result dict (model_info). Spots
    the present keywords and returns the useful definitions."""
    text = " ".join(str(v) for v in model_info.values()).lower() + " " \
           + " ".join(str(k) for k in model_info).lower()
    present = [k for k in GLOSSARY if any(tok in text for tok in k.split())]
    # always include the primary inference if a test ran
    for must in ("randomization test", "p-value"):
        if must not in present:
            present.append(must)
    return explain(*present)


# --------------------------------------------------------------------------- #
# Consolidated interpretation guide (included BY DEFAULT in every report)
# --------------------------------------------------------------------------- #
# Each entry: (Term, Definition, {scopes}). Scopes: "rando" (randomization
# tests), "es" (effect sizes / overlap), "mbd" (multiple baseline), "mv"
# (multivariate), "bayes" (Bayesian models). The "reading" block ("lecture",
# thresholds) is ALWAYS included.
_GUIDE = [
    # - Reading / thresholds (always) -
    ("p-value (randomization)", "p=(1+#>=obs)/(1+B). <0.05 = departure beyond assignment chance. With few cases, the minimal p is bounded by the number of assignments -> 'not significant' != 'no effect'.", {"lecture"}),
    ("Tau-U (overlap)", "Bounded -1..1. Benchmarks (Vannest & Ninci 2015): <0.20 negligible . 0.20-0.60 moderate . 0.60-0.80 strong . >0.80 very strong. |Tau-U|>critical Tau-U = beyond the threshold for this case.", {"lecture"}),
    ("NAP (overlap)", "P(B point > A point). 0.5 = chance. Benchmarks (Parker & Vannest 2009): <0.65 weak . 0.66-0.92 medium . >0.92 strong.", {"lecture"}),
    ("Hedges g / Cohen d", "Standardized difference. Cohen benchmarks: 0.2 small . 0.5 medium . 0.8 large.", {"lecture"}),
    ("ICC", "Share of between-case/patient variance (0..1) - Koo & Li 2016: <0.5 weak . 0.5-0.75 moderate . 0.75-0.90 good . >0.90 excellent.", {"lecture"}),
    ("tau (meta heterogeneity)", "BETWEEN-case standard deviation of the TRUE effects (theta_i ~ Normal(mu, tau)), net of within-case noise, in effect units (points). tau~0 = homogeneous effects; large tau = patients genuinely differ. Poorly estimated with few cases.", {"lecture"}),
    ("I2 (heterogeneity)", "PROPORTION of the variability due to true heterogeneity tau2 vs within-case noise: I2=tau2/(tau2+s2_typ) - Higgins & Thompson 2003: 25% weak . 50% moderate . 75% strong. Also depends on case precision -> read WITH tau.", {"lecture"}),
    ("Prediction interval (meta)", "Range of the effect of a NEW patient (mu +/- dispersion tau). WIDER than the HDI of mu (which is only about the mean). If it crosses 0, a future patient may not respond even if mu is credible. = individual prognosis.", {"lecture"}),
    ("pd (Bayesian)", "Probability of direction. NO 'significance' in Bayesian: 0.75 weak . 0.90 moderate . 0.95 strong . 0.975+ very strong. Only declare a direction from ~0.95 on.", {"lecture"}),
    ("HDI 95% (Bayesian)", "Highest density interval: '95% probability that the parameter lies there'. EXCLUDES 0 = clear; INCLUDES 0 = UNCERTAIN (not no effect).", {"lecture"}),
    ("P(es>ROPE)", "Probability of a PRACTICALLY useful effect (beyond the ROPE). >=0.95 decisive . 0.80-0.95 probable . 0.50-0.80 uncertain . <0.50 below threshold. The ROPE (e.g. 0.5) should reflect the minimum clinically important difference.", {"lecture"}),
    ("LEVEL vs TREND", "Key distinction (spontaneous recovery). LEVEL (does NOT remove the trend): raw Tau-U, NAP, diff, b2 jump, BITS level es. TREND-corrected: baseline-corrected Tau-U (Tarlow 2016), b3 slope / MBD 'slope' test, ITS-with-trend (baseline >= 8 points).", {"lecture"}),
    ("Reading level/trend", "Compare raw vs corrected Tau-U: raw high BUT corrected ~0 -> effect is mostly a CONTINUATION of the baseline trend (spontaneous improvement), not the treatment. BOTH high -> robust effect BEYOND the trend. At the group level, the MBD design offset controls the common trend.", {"lecture"}),
    # - Randomization tests -
    ("Randomization test", "SCED primary inference: the assignment/moment is re-drawn under the actually drawable scheme and the observed statistic is compared to that distribution. Exact (no distributional assumption).", {"rando", "mbd"}),
    ("Partial F", "Variance explained by the condition beyond the nuisance (time/unit). Omnibus (2 or k conditions).", {"rando"}),
    ("Freedman-Lane", "Permutation of the RESIDUALS of the reduced model (time as covariate) -> p of the condition effect net of time.", {"rando"}),
    ("Ordered trend (dose)", "For ordered conditions: slope on dose + Jonckheere-Terpstra. To be pre-specified; not validated for alternating designs.", {"rando"}),
    ("Heterogeneity (Q/I2)", "Does the effect vary across patients? (interaction + Cochran Q/I2).", {"rando"}),
    # - Multiple baseline -
    ("MBD procedure", "Randomization scheme (Levin 2017): WW (case permutation) . MB (start-point with replacement) . MB-R (without replacement) . KL . Rev/Rev-M.", {"mbd"}),
    ("START_WINDOW", "Actual drawing window of the introduction moment (e.g. {5..10}). Providing it makes the test EXACT.", {"mbd"}),
    ("ML b2 / b3 / ICC", "Multilevel model: b2 = level jump (net of time); b3 = slope change (cumulative effect); ICC = between-patient variance. MODEL inference, independent of randomization.", {"mbd"}),
    ("BC-SMD (design-comparable d)", "Jump standardized by the BETWEEN-case SD (Hedges-Pustejovsky-Shadish) -> comparable to an RCT d, meta-analyzable. Small if ICC high. Approx (ref. scdhlm).", {"mbd"}),
    ("VAIOR", "Visual aid (Manolov & Vannest 2019): baseline trend +/- projected MAD; B points green/yellow/red. NOT a test.", {"mbd"}),
    # - Multivariate -
    ("PERMANOVA", ">=2 outcomes: tests whether the condition shifts the joint PROFILE (pseudo-F on distances, by permutation; Anderson 2001). Incomplete cases dropped (complete-case).", {"mv"}),
    ("Holm / FDR", "Multiplicity corrections: Holm controls the risk of >=1 false positive (FWER); Benjamini-Hochberg the proportion of false positives (FDR).", {"mv", "rando"}),
    # - Bayesian -
    ("es / level_change", "Standardized effect size (jump)/sigma, oriented (es>0 = improvement). Small-sample correction built in.", {"bayes"}),
    ("rho (AR1)", "Modeled lag-1 autocorrelation (BITS/BUCP), not corrected a posteriori. 0 in the simple model.", {"bayes"}),
    ("CP (BUCP)", "Unknown change-point: posterior concentrated near the boundary = immediate effect (immediacy); diffuse = uncertain/delayed.", {"bayes"}),
    ("es_end / b3 slope (trend)", "CUMULATIVE/late effect: es_end = effect at end of phase (jump + slope.duration); b3 = slope per session. es_end >> es = effect that builds up (e.g. TMS).", {"bayes"}),
    ("Rhat / ESS / divergences", "MCMC diagnostics: Rhat<1.01 required (1.01-1.05 borderline); high ESS = good; divergences>0 -> doubtful result.", {"bayes"}),
    ("BITS/BUCP prerequisites", ">= ~8 points/phase and expected d >= 3 for a reliable estimate (Natesan Batley 2020); otherwise es very uncertain.", {"bayes"}),
    # - Bayesian reporting rules (BARG; Kruschke 2021, Nature Human Behaviour) -
    ("BARG - reporting guide", "Bayesian Analysis Reporting Guidelines (Kruschke 2021, Nat. Hum. Behav.): pre-specify the model, REPORT (1) likelihood + justified priors, (2) MCMC convergence (R-hat<1.01, sufficient ESS, 0 divergences), (3) the POSTERIOR = median + HDI (declared level), (4) the DECISION via HDI + ROPE (and/or pd), (5) a SENSITIVITY analysis to the priors. Pre-specify, do not adjust a posteriori.", {"bayes"}),
    ("BARG - decision (HDI + ROPE)", "Conclude a CREDIBLE effect = the 95% HDI EXCLUDES the ROPE (Kruschke 2018) - more demanding than 'excludes 0'. ROPE = practical equivalence interval = MCID (minimum clinically important difference), to be pre-specified. % of the HDI in the ROPE: <2.5% -> reject equivalence (effect); >97.5% -> accept equivalence (no useful effect); in between -> undecided (more data).", {"bayes"}),
    ("BARG - pd vs magnitude", "The pd (probability of direction) measures EXISTENCE/sign, NOT size nor clinical usefulness. Correspondence (Makowski 2019): pd 0.95~p.10 . 0.975~p.05 . 0.99~p.02. Never conclude on pd alone -> always HDI (magnitude) + ROPE (relevance). HDI level (95%/89%) = convention to pre-specify and justify (89% McElreath: more stable tails; 95% expected in clinical work).", {"bayes"}),
]


def _ascii_plain(s):
    """Replace 'fancy' characters with ASCII equivalents (plain reports)."""
    repl = {"▶ ": "", "▶": "", "▸ ": "", "▸": "", "→": "->", "←": "<-",
            "≥": ">=", "≤": "<=",
            "×": "x", "·": ".", "•": "-", "●": "-", "◆": "-", "■": "-",
            "²": "2", "³": "3",
            "\u2014": "-", "\u2013": "-", "\u2011": "-", "’": "'", "‘": "'",
            "œ": "oe", "æ": "ae", "≈": "~",
            "…": "...", "«": '"', "»": '"', "“": '"', "”": '"',
            "√": "sqrt", "∏": "prod",
            "∑": "sum", "±": "+/-", "≠": "!=", "∞": "inf", "°": "deg",
            "μ": "mu", "σ": "sigma", "τ": "tau", "ρ": "rho", "ν": "nu",
            "φ": "phi",
            "Δ": "Delta", "α": "alpha", "β": "beta", "χ": "chi"}
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


def ascii_sanitize_df(df):
    """Return a COPY of the DataFrame where every 'fancy' character (headers AND text
    cells) is replaced by its ASCII equivalent. To be applied to EACH sheet before xlsx
    writing to guarantee reports without special characters."""
    import pandas as pd
    out = df.copy()
    out.columns = [_ascii_plain(str(c)) for c in out.columns]
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = out[col].map(lambda v: _ascii_plain(v) if isinstance(v, str) else v)
    return out


def dedup_pooled_rows(df, unit_col, *, group_col="Group", pooled_label="(all)", tol=0.01):
    """Anti-redundancy: if a unit's row (case/tier) is IDENTICAL (up to ``tol`` rounding)
    between its cohort and the pooled group ('(all)'), keep only the cohort; otherwise keep
    both (e.g. Bayesian shrinkage that differs by pooling). For per-unit statistics that are
    independent (Tau-U, NAP...) and identical by construction -> the redundant pooled version
    is dropped."""
    import numpy as np
    import pandas as pd
    if df.empty or group_col not in df.columns or unit_col not in df.columns:
        return df
    idcols = [c for c in (group_col, "Outcome", unit_col) if c in df.columns]
    num = [c for c in df.columns if c not in idcols and df[c].dtype != object]
    keycols = [c for c in ("Outcome", unit_col) if c in df.columns]
    keep = []
    for _, sub in df.groupby(keycols, dropna=False):
        pooled = sub[sub[group_col] == pooled_label]
        coh = sub[sub[group_col] != pooled_label]
        if len(pooled) and len(coh):
            same = bool(num) and np.allclose(pooled[num].iloc[0].values, coh[num].iloc[0].values,
                                             rtol=0, atol=tol, equal_nan=True)
            keep.append(coh if same else sub)
        else:
            keep.append(sub)
    return pd.concat(keep, ignore_index=True)


def data_recap_df(df, *, tier_col, session_col=None, phase_col=None, outcomes=None,
                  group_col=None, baseline=None, treatment=None, design=None, extra=None):
    """First sheet of any report: recap of the DATA & DESIGN (ASCII). Describes: design,
    analyzed outcomes, included variables, number of units, number of observations, phases,
    sessions/unit, cohorts. ``extra`` = list of (key, value) to append."""
    import numpy as np
    import pandas as pd
    has_tier = bool(tier_col) and tier_col in df
    rows = []
    if design:
        rows.append(("Design", design))
    if outcomes is not None:
        outs = outcomes if isinstance(outcomes, (list, tuple)) else [outcomes]
        rows.append(("Analyzed outcome(s)", ", ".join(map(str, outs)) + f"  (n={len(outs)})"))
    rows.append(("N units (tiers/cases)", int(df[tier_col].nunique()) if has_tier else 1))
    rows.append(("N observations", int(len(df))))
    if group_col and group_col in df:
        gg = df.groupby(group_col)[tier_col].nunique()
        rows.append(("Cohorts (" + str(group_col) + ")",
                     ", ".join(f"{k}={v}" for k, v in gg.items())))
    if phase_col and phase_col in df:
        ph = [str(p) for p in pd.unique(df[phase_col].dropna())]
        rows.append(("Observed phases", ", ".join(ph)))
    if baseline is not None or treatment is not None:
        rows.append(("Contrast", f"baseline={baseline} -> treatment={treatment}"))
    if session_col and session_col in df and has_tier:
        per = df.groupby(tier_col)[session_col].count()
        rows.append(("Sessions / unit (min-med-max)",
                     f"{int(per.min())}-{int(per.median())}-{int(per.max())}"))
    elif session_col and session_col in df:
        rows.append(("N sessions", int(df[session_col].notna().sum())))
    cols_struct = [c for c in (tier_col, session_col, phase_col, group_col) if c]
    measure_cols = outcomes if isinstance(outcomes, (list, tuple)) else \
        [c for c in df.columns if c not in cols_struct]
    rows.append(("Included variables", ", ".join(map(str, list(cols_struct) + list(measure_cols)))))
    for k, v in (extra or []):
        rows.append((str(k), v))
    return pd.DataFrame(rows, columns=["Element", "Value"])


def write_report(path, sheets, *, mode="w", if_sheet_exists="replace"):
    """Write a .xlsx with GUARANTEED ASCII on each sheet (headers + cells). ``sheets`` =
    ordered dict {sheet_name: DataFrame}. ``mode='a'`` appends to an existing file. Silent
    if the file is locked/open (does not interrupt the analysis). Returns the path or None."""
    import pandas as pd
    try:
        kw = {"mode": "a", "if_sheet_exists": if_sheet_exists} if mode == "a" else {}
        with pd.ExcelWriter(path, engine="openpyxl", **kw) as w:
            for name, df in sheets.items():
                if df is None:
                    continue
                ascii_sanitize_df(df).to_excel(w, sheet_name=str(name)[:31], index=False)
        return path
    except Exception:
        return None


def write_stacked(writer, sheet_name, blocks, *, sanitize=True, styles=None):
    """Write SEVERAL tables on ONE sheet, stacked vertically, each preceded by a title.
    ``blocks`` = ordered list of ``(title, DataFrame)``. ``styles`` (optional) = list parallel
    to ``blocks``; each element None or dict ``{"rules":..., "row_rules":...}`` applied to THIS
    block (cf. ``style_cells``) accounting for the row offset. Readability: auto widths (capped),
    bold titles + headers, wrap on long cells. ASCII guaranteed if ``sanitize``. Returns writer."""
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    name = str(sheet_name)[:31]
    start = 0
    widths = {}
    ws = None
    for bi, (title, dfb) in enumerate(blocks):
        if dfb is None:
            continue
        d = ascii_sanitize_df(dfb) if sanitize else dfb
        d.to_excel(writer, sheet_name=name, startrow=start + 1, index=False)   # title on the row above
        ws = writer.sheets[name]
        tcell = ws.cell(row=start + 1, column=1, value=_ascii_plain(str(title)))
        tcell.font = Font(bold=True, size=12)
        for c in range(1, len(d.columns) + 1):                                 # bold header
            ws.cell(row=start + 2, column=c).font = Font(bold=True)
        spec = (styles or [None] * len(blocks))[bi]                            # coloring of this block
        if spec:
            style_cells(ws, d, spec.get("rules", {}), header_rows=start + 2,
                        row_rules=spec.get("row_rules"))
        for j, col in enumerate(d.columns):                                    # track widths (per column)
            cells = [str(col)] + [("" if v is None else str(v)) for v in d.iloc[:, j].tolist()]
            w = max(len(x) for x in cells)
            if j == 0:
                w = max(w, len(str(title)))
            widths[j] = max(widths.get(j, 0), w)
        start += len(d) + 3                                                    # title + header + data + blank row
    if ws is not None:                                                          # widths (capped) + wrap wide columns
        for j, w in widths.items():
            col = get_column_letter(j + 1)
            ws.column_dimensions[col].width = min(max(w + 2, 12), 70)
            if w + 2 > 70:                                                      # long text -> wrap
                for cell in ws[col]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return writer


def style_cells(ws, df, rules, *, header_rows=1, row_rules=None, col_offset=0):
    """Highlight cells of a sheet already written by ``df.to_excel``.
    ``rules`` = {column_name: fn(value) -> (bold, hex_color|None)} (judges the cell alone).
    ``row_rules`` = {column_name: fn(row_Series) -> (bold, hex_color|None)} (judges from the
    WHOLE row: useful to color an HDI by its two bounds). ``header_rows`` = Excel row
    (1-indexed) of the header; ``col_offset`` = column offset. Color code: green C6EFCE
    (credible/strong), amber FFE699, yellow FFF2CC, gray F2F2F2, RED FFC7CE (problem).
    Silent per cell."""
    from openpyxl.styles import Font, PatternFill
    cols = list(df.columns)
    row_rules = row_rules or {}
    for cj, col in enumerate(cols):
        fn = rules.get(col); rfn = row_rules.get(col)
        if fn is None and rfn is None:
            continue
        for ri in range(len(df)):
            try:
                bold, fill = (rfn(df.iloc[ri]) if rfn is not None else fn(df.iloc[ri, cj]))
            except Exception:
                bold, fill = False, None
            cell = ws.cell(row=header_rows + 1 + ri, column=cj + 1 + col_offset)
            if bold:
                cell.font = Font(bold=True)
            if fill:
                cell.fill = PatternFill(fgColor=fill, fill_type="solid")
    return ws


def _to_float(v):
    """Coerce a value to float, returning None on failure (instead of raising)."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def hdi_excludes_zero(lo, hi):
    """True if the interval [lo, hi] does not cross 0 (same signs, finite bounds)."""
    lo, hi = _to_float(lo), _to_float(hi)
    if lo is None or hi is None:
        return False
    return (lo > 0 and hi > 0) or (lo < 0 and hi < 0)


def hdi_str_fill(s):
    """(bold, color) for a text-format HDI cell '[lo;hi]': GREEN if it EXCLUDES 0 (effect
    credible in one direction), otherwise light gray. Tolerates ',' or ';' as separator."""
    import re
    m = re.findall(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", str(s))
    if len(m) < 2:
        return (False, None)
    return (True, "C6EFCE") if hdi_excludes_zero(m[0], m[-1]) else (False, "F2F2F2")


def rhat_fill(v):
    """RED if R-hat > 1.01 (unconverged chains), amber if > 1.005, otherwise nothing."""
    a = _to_float(v)
    if a is None:
        return (False, None)
    if a > 1.01:
        return (True, "FFC7CE")
    return (False, "FFE699") if a > 1.005 else (False, None)


def ess_fill(v):
    """RED if ESS < 100 (unreliable estimate), amber if < 400 (borderline), otherwise nothing."""
    a = _to_float(v)
    if a is None:
        return (False, None)
    if a < 100:
        return (True, "FFC7CE")
    return (False, "FFE699") if a < 400 else (False, None)


def param_table_styles():
    """Style rules for the 'Parameters (full)' sheet (output of ``full_param_table``).
    Returns ``(rules, row_rules)``: HDI (hdi_3%/hdi_97%) colored GREEN if the interval
    EXCLUDES 0 (judged on the row's two bounds); r_hat / ess_bulk / ess_tail RED if doubtful."""
    def _hdi_row(row):
        lo = row.get("hdi_3%"); hi = row.get("hdi_97%")
        return (True, "C6EFCE") if hdi_excludes_zero(lo, hi) else (False, None)
    rules = {"r_hat": rhat_fill, "ess_bulk": ess_fill, "ess_tail": ess_fill}
    row_rules = {"hdi_3%": _hdi_row, "hdi_97%": _hdi_row}
    return rules, row_rules


def cohen_fill(g, *, bands=(0.2, 0.5, 0.8)):
    """(bold, color) for a Cohen-type SMD: negligible gray / small yellow / medium amber /
    large green (+ bold from 'medium' on). ``bands`` adaptable (e.g. Tau-U: (0.20, 0.60, 0.80))."""
    try:
        a = abs(float(g))
    except (TypeError, ValueError):
        return (False, None)
    lo, mid, hi = bands
    if a < lo:
        return (False, "F2F2F2")
    if a < mid:
        return (False, "FFF2CC")
    if a < hi:
        return (True, "FFE699")
    return (True, "C6EFCE")


def bayesian_diag_legend():
    """PRECISE interpretation guide for each MCMC diagnostic plot (cf. plot_bayesian_diag /
    plot_bayesian_panel). Returns a DataFrame (Plot, How to read, Warning sign). Written next
    to the figures so each diagnostic is readable without expertise."""
    import pandas as pd
    rows = [
        ("trace", "Each MCMC chain over the iterations (+ density on the left).",
         "Good: overlaid, stationary chains ('fuzzy caterpillar'). Alarm: separated chains, drift, jumps."),
        ("rank", "RANK histograms per chain (more sensitive than the trace).",
         "Good: ~uniform/flat bars across chains. Alarm: U shape / staircase = poor mixing."),
        ("forest", "Effects side by side: point = median, bar = 95% HDI, + r_hat and ess shown.",
         "Read size/sign and whether the HDI excludes 0. Alarm: r_hat > 1.01 or low ess next to the name."),
        ("posterior", "Posterior density of the effect + 95% HDI (+ ROPE if provided).",
         "Read median, HDI width, position vs 0/ROPE. HDI including 0 = uncertainty, NOT absence of effect."),
        ("energy", "NUTS-specific diagnostic (BFMI): marginal energy vs transition energy.",
         "Good: the two histograms overlap. Alarm: offset = low BFMI, inefficient exploration (reparameterize)."),
        ("ess", "Evolution of the EFFECTIVE sample size with the number of draws.",
         "Good: increases, ESS > 400 per parameter. Alarm: low plateau = chains too correlated."),
        ("autocorr", "Autocorrelation of the draws, per parameter and chain.",
         "Good: fast drop to 0. Alarm: slow decay = strong autocorrelation (little info per draw)."),
        ("qq_loo_pit", "LOO-PIT calibration (analogue of a QQ-plot): cross-validated PIT vs uniform.",
         "Good: ECDF curve within the confidence band. Alarm: S shape / out of band = poor calibration "
         "(family or variance mis-specified)."),
        ("ppc", "Posterior predictive check: density of the OBSERVED data vs model replicates.",
         "Good: the observed falls within the replicate spread. Alarm: systematically offset = poor fit."),
        ("ppc_by_phase / ppc_by_case (grouped/ folder)", "GROUPED PPC: same reading but PER level "
         "(phase A vs B, or per patient) -> much sharper than the pooled marginal PPC.",
         "Spot a specific GROUP where the observed leaves the spread (local misfit); the marginal hides it."),
        ("panel", "Dashboard: forest + rank + energy + PPC (or posterior) in one figure.",
         "Quick overview; refer to the dedicated plot for the detail."),
    ]
    return pd.DataFrame(rows, columns=["Plot", "How to read", "Warning sign / interpretation"])


def full_param_table(models, *, drop_helpers=True):
    """EXHAUSTIVE table of ALL posterior parameters of each fitted model (mean, sd, 94% HDI,
    mcse, ess_bulk/tail, r_hat) via ``arviz.summary``. ``models`` = dict ``{key: InferenceData}``
    (str or tuple key). ``drop_helpers`` removes the non-interpretable auxiliary variables
    (non-centered ``z_*``, internal ``chol_re``). Feeds the secondary sheet 'Parameters (full)':
    nothing is hidden. Empty DataFrame if none."""
    import pandas as pd
    blocks = []
    for key, idata in (models or {}).items():
        if idata is None:
            continue
        try:
            import arviz as az
            sm = az.summary(idata)
        except Exception:
            continue
        sm = sm.reset_index().rename(columns={"index": "parameter"})
        if drop_helpers:
            sm = sm[~sm["parameter"].astype(str).str.startswith(("z_", "chol_re"))]
        lab = key if isinstance(key, str) else " / ".join(str(k) for k in key)
        sm.insert(0, "Model", lab)
        blocks.append(sm)
    return pd.concat(blocks, ignore_index=True) if blocks else pd.DataFrame()


def save_idata(idata, path, name):
    """Save a fitted Bayesian model (ArviZ ``InferenceData``) in netCDF, to reload it without
    re-running the MCMC (``arviz.from_netcdf``). Returns the written path or None. Silent if
    idata is None or if writing fails."""
    import os
    if idata is None:
        return None
    try:
        os.makedirs(path, exist_ok=True)
        safe = "".join(c if (c.isalnum() or c in "-_") else "_" for c in str(name))
        fp = os.path.join(path, f"{safe}.nc")
        idata.to_netcdf(fp)
        return fp
    except Exception:
        return None


def interpretation_glossary(scope="all", plain=True):
    """
    Consolidated interpretation glossary as a ``DataFrame`` (Term / Definition), meant to be
    added as an 'Interpretation guide' sheet of each report. ``scope`` filters the relevant
    sections: ``"rando"``, ``"mbd"``, ``"mv"``, ``"bayes"`` (or a list), ``"all"`` = everything.
    The READING block (pd/HDI/ROPE/Tau-U/NAP/ICC/I2 thresholds) is ALWAYS included.
    ``plain=True`` returns an ASCII version (without bullets or fancy symbols).
    """
    import pandas as pd
    if scope == "all":
        want = {"lecture", "rando", "es", "mbd", "mv", "bayes"}
    else:
        want = {"lecture"} | (set(scope) if not isinstance(scope, str) else {scope})
    rows = [(t, d) for (t, d, sc) in _GUIDE if sc & want]
    if plain:
        rows = [(_ascii_plain(t), _ascii_plain(d)) for t, d in rows]
    return pd.DataFrame(rows, columns=["Term", "Definition"])


def append_glossary_sheet(xlsx_path, scope="all", sheet_name="Interpretation guide"):
    """Add (or replace) the glossary sheet to an EXISTING .xlsx report. Idempotent; silent if
    the file is locked/absent."""
    import os
    import pandas as pd
    if not xlsx_path or not os.path.exists(xlsx_path):
        return None
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as w:
            interpretation_glossary(scope).to_excel(w, sheet_name=sheet_name, index=False)
        return xlsx_path
    except Exception:
        return None


def provenance_df(*, seed=None, extra=None):
    """PROVENANCE block (reproducibility): package versions, seed, timestamp. To stack as the
    first sheet of each report (BARG / good practice)."""
    import sys
    import datetime
    import importlib
    import pandas as pd
    rows = [("Analysis date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Python", sys.version.split()[0])]
    for pkg in ("numpy", "scipy", "pandas", "statsmodels", "pymc", "arviz"):
        try:
            rows.append((pkg, importlib.import_module(pkg).__version__))
        except Exception:
            rows.append((pkg, "absent"))
    rows.append(("BC-SMD", "R scdhlm (REML) -- see the R environment"))
    if seed is not None:
        rows.append(("Seed", seed))
    for k, v in (extra or []):
        rows.append((str(k), v))
    return pd.DataFrame(rows, columns=["Element", "Value"])


def data_quality_df(df, *, tier_col, session_col, phase_col, outcome_col,
                    baseline=None, treatment=None):
    """DATA QUALITY block per case (SCRIBE 2016): n per phase, missing data (non-numeric/NaN
    outcome) and largest session gap. Advisory."""
    import numpy as np
    import pandas as pd
    rows = []
    for t, g in df.groupby(tier_col):
        g = g.sort_values(session_col)
        y = pd.to_numeric(g[outcome_col], errors="coerce")
        ph = g[phase_col].astype(str) if phase_col and phase_col in g else pd.Series([], dtype=str)
        nA = int((ph == str(baseline)).sum()) if (baseline is not None and len(ph)) else np.nan
        nB = int((ph == str(treatment)).sum()) if (treatment is not None and len(ph)) else np.nan
        miss = int(y.isna().sum())
        pct = round(100.0 * miss / max(len(g), 1), 1)
        s = pd.to_numeric(g[session_col], errors="coerce").dropna().to_numpy()
        gap = int(np.max(np.diff(np.sort(s)))) if len(s) > 1 else 0
        rows.append({"Case": t, "n total": int(len(g)), "n_A": nA, "n_B": nB,
                     "n missing": miss, "% missing": pct, "max session gap": gap})
    return pd.DataFrame(rows)


def slug_token(s):
    """Clean ASCII token for a folder/file name: alnum/_/- kept, the rest -> '_'.
    None/'(all)'/'' -> 'all'. Used for outcomes and cohorts in the directory tree."""
    if s is None or str(s).strip() in ("(all)", "all", ""):
        return "all"
    return "".join(c if (c.isalnum() or c in "-_") else "_" for c in str(s))


def std_layout(output_dir, outcome=None, analysis=None):
    """STANDARD output directory tree (built into every SCED pipeline), ORGANIZED BY OUTCOME.

        <output_dir>/<outcome>/
          Plot/                    descriptive (tiered, panels, VAIOR, Brinley) - model-independent
          Plot/poole/              pooled inferential + Bayesian (hier & meta) - model-dependent
          Plot/forest/<model>/     forests - model-dependent
          Analyse/<analysis>/      .xlsx reports (+ models/ + diagnostics/ for the Bayesian side)

    ``analysis`` names the analysis SUB-FOLDER, e.g.:
        'permutation_test'  (formerly 'MBD', randomization test + multilevel)
        'bayes/hier'        (hierarchical partial: comparison + per-model reports + cache)
        'bayes/meta'        (two-stage meta)
        'phase_design' | 'alternating' | 'bayes/condition'

    ``models`` and ``diagnostics`` live INSIDE the analysis folder (no global models/).
    Returns a dict of paths (not created; created on write)."""
    import os
    oc = slug_token(outcome) if outcome is not None else "_"
    base = os.path.join(output_dir, oc)
    plot = os.path.join(base, "Plot")
    analyse = os.path.join(base, "Analyse", analysis) if analysis else os.path.join(base, "Analyse")
    return {"analyse": analyse, "plot": plot,
            "pooled": os.path.join(plot, "poole"), "forest": os.path.join(plot, "forest"),
            "models": os.path.join(analyse, "models"), "diagnostics": os.path.join(analyse, "diagnostics")}
